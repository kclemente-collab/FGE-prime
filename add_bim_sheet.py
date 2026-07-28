#!/usr/bin/env python3
"""
Adds FGE-OS CAMPUS BIM SPECIFICATION v1.0 sheet to the existing constitutional matrix.
Production Mode | Single Source of Truth
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = load_workbook('/home/workdir/artifacts/FGE_Character_Material_Fingerprint_Matrix_v1.0_Phase1_Core8.xlsx')

# Styling (matching existing file)
fill_black = PatternFill(start_color="0D0D0D", end_color="0D0D0D", fill_type="solid")
fill_navy = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
fill_gold = PatternFill(start_color="C9A227", end_color="C9A227", fill_type="solid")
fill_dark_green = PatternFill(start_color="0D3D0D", end_color="0D3D0D", fill_type="solid")
fill_crimson = PatternFill(start_color="4A0E0E", end_color="4A0E0E", fill_type="solid")

font_title = Font(name="Arial", size=18, bold=True, color="C9A227")
font_header = Font(name="Arial", size=11, bold=True, color="000000")
font_subheader = Font(name="Arial", size=12, bold=True, color="C9A227")
font_body = Font(name="Arial", size=9, color="E0E0E0")
font_locked = Font(name="Arial", size=9, bold=True, color="00FF00")
font_small = Font(name="Arial", size=8, color="AAAAAA")

thin_border = Border(
    left=Side(style='thin', color='C9A227'),
    right=Side(style='thin', color='C9A227'),
    top=Side(style='thin', color='C9A227'),
    bottom=Side(style='thin', color='C9A227')
)

# Create new sheet
ws = wb.create_sheet("04_FGEOS_Campus_BIM_Spec_v1.0")

# ========== TITLE BLOCK ==========
ws.merge_cells('A1:L2')
ws['A1'] = "FGE-OS CAMPUS BIM SPECIFICATION v1.0"
ws['A1'].font = font_title
ws['A1'].fill = fill_black
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('A3:L3')
ws['A3'] = "Physical-Digital Operating Architecture | Single Source of Truth | Integrated with FGE Character Material Fingerprint Matrix v1.0 & FGE-OS Master Campus Blueprint v1.0"
ws['A3'].font = font_subheader
ws['A3'].fill = fill_navy
ws['A3'].alignment = Alignment(horizontal='center')

ws.merge_cells('A5:L5')
ws['A5'] = "PURPOSE: Transform the FGE Master Campus from visual concept into a fully instantiable, machine-readable, canon-locked technical specification. Every node, connector, room, and vertical system is now addressable and production-ready."
ws['A5'].font = font_body
ws['A5'].fill = fill_navy
ws['A5'].alignment = Alignment(horizontal='left', wrap_text=True)
ws.row_dimensions[5].height = 35

# ========== SECTION 1: NODE ID STANDARD ==========
ws.merge_cells('A7:L7')
ws['A7'] = "SECTION 1 — NODE ID STANDARD (Locked)"
ws['A7'].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ws['A7'].fill = fill_gold
ws['A7'].alignment = Alignment(horizontal='center')

ws.merge_cells('A8:L8')
ws['A8'] = "Format: FGE-[FLOOR/ORGAN]-[NUMBER]-[TYPE]   |   Example: FGE-F2-030-QC = Floor 2, Node 030, Quality Control type"
ws['A8'].font = font_body
ws['A8'].fill = fill_navy
ws['A8'].alignment = Alignment(horizontal='left')

# ========== SECTION 2: MAIN NODE MAPPING TABLE ==========
ws.merge_cells('A10:L10')
ws['A10'] = "SECTION 2 — CAMPUS NODE MAPPING MATRIX (Core Nodes v1.0)"
ws['A10'].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ws['A10'].fill = fill_gold
ws['A10'].alignment = Alignment(horizontal='center')

# Headers
headers = ["Node ID", "Floor / Organ", "Name", "Primary Function", "Visual Signature", "Connectors", "Status", "Key Connections", "Coordinates (Relative)", "Notes"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=11, column=col, value=header)
    cell.font = font_header
    cell.fill = fill_gold
    cell.font = Font(name="Arial", size=8, bold=True, color="000000")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws.row_dimensions[11].height = 22

# Node data (core nodes from blueprint + Floor 2 System Vault Organ)
nodes = [
    ["FGE-F0-001-FOUND", "Foundation", "Memory Vault Core", "Knowledge Infrastructure & Archive", "Dark marble + gold Kintsugi + subtle glow", "POWER, DATA, ARCHIVE", "Canonized", "FGE-F1-020, FGE-F2-011", "Base Layer", "Historian / Librarian Agents"],
    ["FGE-F1-020-CANON", "Floor 1 - Canon Genome", "Canon Genome Interface", "Identity Preservation & Drift Prevention", "Obsidian arch + glowing DNA helix", "DATA, CANON, QC", "Canonized", "FGE-F0-001, FGE-F2-030", "Level 1", "Curator + Canon Guardian Agents"],
    ["FGE-F2-011-ASSET", "Floor 2 - System Vault", "Asset Passport System", "Permanent identity + lineage tracking", "Dark marble + gold filigree badge", "DATA, QC, INV", "Canonized", "FGE-F2-020, FGE-F2-030", "Level 2 East", "Core of Gated Production Container"],
    ["FGE-F2-020-CANON", "Floor 2 - System Vault", "Canon Genome Interface (Vault)", "Prevent identity drift on all assets", "Obsidian arch + glowing DNA helix", "DATA, CANON, QC", "Canonized", "FGE-F1-020, FGE-F2-011", "Level 2 Center", "Links to Floor 1 Canon"],
    ["FGE-F2-030-QC", "Floor 2 - System Vault", "QC Intelligence Engine", "Immune system & worthiness scoring", "Clinical white + gold measurement tools", "DATA, QC, TASK, MAVEN", "Canonized", "FGE-F2-011, FGE-F2-040, FGE-F2-050", "Level 2 West", "Quality Gate for all production"],
    ["FGE-F2-040-INV", "Floor 2 - System Vault", "Inventory State Machine", "Controls all movement via tasks only", "Heavy vault door with state indicators", "TASK, INV, MAVEN", "Canonized", "FGE-F2-030, FGE-F2-051", "Level 2 Core", "No direct sales access"],
    ["FGE-F2-050-MAVEN", "Floor 2 - System Vault", "Maven of Product Ascension", "Refinement & optimization intelligence", "Polished obsidian desk with tools", "DATA, TASK, QC, FEEDBACK", "Canonized", "FGE-F2-030, FGE-F2-051, FGE-F5-ECON", "Level 2 West Wing", "Key agent for continuous improvement"],
    ["FGE-F2-051-TASK", "Floor 2 - System Vault", "Task Dispatcher", "Controls all movement through tasks", "Brass elevator panel with tickets", "TASK, ELEV, ALL F2", "Canonized", "All F2 nodes + Vertical Spine", "Level 2 Central Hub", "Heart of gated production flow"],
    ["FGE-F3-030-QUAL", "Floor 3 - Quality", "Quality Intelligence Labs", "Judgment, scoring & approval gate", "Clean technical + gold measurement", "DATA, QC, EVO", "Canonized", "FGE-F2-030, FGE-F4-EVO", "Level 3", "Critic + QA Agents"],
    ["FGE-F4-EVO-001", "Floor 4 - Evolution", "Evolution Reactor Core", "Mutation, testing & inheritance", "Organic crystalline + neural glow", "DATA, EVO, MUTATION", "Canonized", "FGE-F3-030, FGE-F2-050", "Level 4", "Scientist + Evolution Agents"],
    ["FGE-F5-ECON-001", "Floor 5 - Economic", "Economic Engine Core", "Value creation & revenue flow", "Luxury dark wood + gold brass", "DATA, REV, MARKET", "Canonized", "FGE-F2-050, FGE-F6-INTEL", "Level 5", "Market + Finance + Licensing Agents"],
    ["FGE-F6-INTEL-001", "Floor 6 - Intelligence", "Intelligence Command Core", "Strategy, prediction & resource allocation", "Dark neural chamber + floating data", "DATA, STRAT, ALL", "Canonized", "FGE-F5-ECON, Penthouse", "Level 6", "Architect + Strategist Agents"],
    ["FGE-PENT-001-CIV", "Penthouse - Civilization", "Constitution Core / Throne Room", "Final governance & mission alignment", "Circular command + panoramic glass + genome displays", "GOV, FINAL, ALL", "Canonized", "FGE-F6-INTEL", "Crown Level", "Executive + Constitution Guardian Agents"],
]

for row_idx, node in enumerate(nodes, 12):
    for col_idx, value in enumerate(node, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = font_body
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = thin_border
        
        if col_idx == 7:  # Status column
            cell.fill = fill_dark_green
            cell.font = font_locked
        if col_idx == 1:  # Node ID
            cell.font = Font(name="Arial", size=8, bold=True, color="C9A227")

# Set column widths
col_widths = [18, 18, 26, 32, 38, 28, 12, 32, 16, 28]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

for row in range(12, 24):
    ws.row_dimensions[row].height = 38

# ========== SECTION 3: CONNECTOR DEFINITIONS ==========
ws.merge_cells('A25:L25')
ws['A25'] = "SECTION 3 — CONNECTOR / UTILITY PORT DEFINITIONS (Standard Building Interfaces)"
ws['A25'].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ws['A25'].fill = fill_gold
ws['A25'].alignment = Alignment(horizontal='center')

connector_headers = ["Connector Type", "Purpose", "What it Accepts", "What it Returns", "Security Level", "Example Use"]
for col, header in enumerate(connector_headers, 1):
    cell = ws.cell(row=26, column=col, value=header)
    cell.font = font_header
    cell.fill = fill_gold
    cell.font = Font(name="Arial", size=8, bold=True, color="000000")
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

connectors = [
    ["POWER", "Computational & operational resources", "CPU, Storage, Render capacity, Budget", "Allocated resources + usage logs", "High", "Render Agent requests render budget"],
    ["DATA", "Information highway", "Images, Documents, Telemetry, Market data, Genome", "Processed data, Reports, Decisions", "Medium-High", "QC Engine pulls asset metrics"],
    ["AGENT", "Agent deployment & communication", "Agent role, Mission, Permissions", "Agent actions, Reports, Status", "High", "Maven Agent assigned to Floor 2"],
    ["PRODUCTION", "Creation & modification pipeline", "Asset genome, Character, Scene", "Finished artifact, Product, Package", "Medium", "Task Dispatcher routes to Manufacturing"],
    ["SECURITY", "Permission & audit layer", "Read/Write/Mutate/Export requests", "Approved/Denied + audit trail", "Critical", "Penthouse final approval gate"],
    ["EVOLUTION", "Mutation & inheritance channel", "Current version + proposed change", "Approved mutation + inheritance record", "High", "Floor 4 Evolution Reactor"],
]

for row_idx, conn in enumerate(connectors, 27):
    for col_idx, value in enumerate(conn, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = font_body
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = thin_border

for row in range(27, 33):
    ws.row_dimensions[row].height = 32

# ========== SECTION 4: VERTICAL INFRASTRUCTURE ==========
ws.merge_cells('A34:L34')
ws['A34'] = "SECTION 4 — VERTICAL INFRASTRUCTURE (The Building's Nervous & Circulatory Systems)"
ws['A34'].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ws['A34'].fill = fill_gold
ws['A34'].alignment = Alignment(horizontal='center')

ws.merge_cells('A35:L36')
ws['A35'] = """DATA ELEVATORS (Upward Information Flow): Foundation → Canon → Quality → Evolution → Intelligence → Penthouse
PRODUCTION CONVEYOR (Downward Creation Flow): Strategy → Canon Validation → Manufacturing → Quality → Product → Market
EVOLUTION NERVOUS SYSTEM (Lateral Learning): Every successful mutation on any floor becomes available campus-wide via Floor 4 Evolution Reactor"""
ws['A35'].font = font_body
ws['A35'].fill = fill_navy
ws['A35'].alignment = Alignment(horizontal='left', wrap_text=True)
ws.row_dimensions[35].height = 45
ws.row_dimensions[36].height = 20

# ========== SECTION 5: GATED PRODUCTION + FLOOR 2 INTEGRATION ==========
ws.merge_cells('A38:L38')
ws['A38'] = "SECTION 5 — INTEGRATION: GATED PRODUCTION CONTAINER + FLOOR 2 SYSTEM VAULT ORGAN"
ws['A38'].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ws['A38'].fill = fill_crimson
ws['A38'].alignment = Alignment(horizontal='center')

ws.merge_cells('A39:L41')
ws['A39'] = """Floor 2 (System Vault Organ) is the primary Gated Production Container for the entire Campus.
Core Principle: Nothing reaches Sales except through Tasks. The Vault governs product readiness, not creation.
Key Nodes: FGE-F2-011-ASSET (Passport) → FGE-F2-030-QC (Scoring) → FGE-F2-040-INV (State Machine) → FGE-F2-051-TASK (Dispatcher) → FGE-F2-050-MAVEN (Optimization)
Maven of Product Ascension is the dedicated refinement agent. All weak metrics become improvement tasks. Sales feedback loops back into Floor 2 for continuous inheritance.
This implements the "Observance Level 1" self-awareness of the building: it can identify, classify, score, track lineage, approve movement, and learn from outcomes."""
ws['A39'].font = font_body
ws['A39'].fill = fill_navy
ws['A39'].alignment = Alignment(horizontal='left', wrap_text=True)
ws.row_dimensions[39].height = 75

# ========== SECTION 6: VERSION & CANON LOCK ==========
ws.merge_cells('A43:L43')
ws['A43'] = "SECTION 6 — VERSION CONTROL & CANON LOCK"
ws['A43'].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ws['A43'].fill = fill_gold
ws['A43'].alignment = Alignment(horizontal='center')

ws.merge_cells('A44:L45')
ws['A44'] = f"""Version: v1.0 | Date: {datetime.now().strftime('%Y-%m-%d')}
Status: CANONIZED | Drift Check: ZERO
This specification is now part of the single constitutional source of truth.
All future Campus renders, agent deployments, task routing, and tenant installations must reference Node IDs and Connectors defined herein.
Next planned expansion: v1.1 — Full Floor Plans with exact room coordinates + BIM object library."""
ws['A44'].font = font_body
ws['A44'].fill = fill_navy
ws['A44'].alignment = Alignment(horizontal='left', wrap_text=True)
ws.row_dimensions[44].height = 50

# Final note
ws.merge_cells('A47:L47')
ws['A47'] = "FGE-OS CAMPUS BIM SPECIFICATION v1.0 — The building is now a permanent, portable, instantiable operating architecture. Tenants evolve. The skeleton remains."
ws['A47'].font = Font(name="Arial", size=9, italic=True, color="C9A227")
ws['A47'].fill = fill_black
ws['A47'].alignment = Alignment(horizontal='center')

# Freeze panes
ws.freeze_panes = 'A11'

# Save
wb.save('/home/workdir/artifacts/FGE_Character_Material_Fingerprint_Matrix_v1.0_Phase1_Core8.xlsx')
print("BIM Specification sheet added successfully.")
print("New sheet: 04_FGEOS_Campus_BIM_Spec_v1.0")