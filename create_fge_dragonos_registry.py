#!/usr/bin/env python3
"""
FGE DragonOS v2.1 — Living Architecture Registry + Business Maven Launch Kit
Collector-Grade XLSX | FGE Publishing Studio Constitution Compliant
Every asset connects to Character | Companion | Location | Event | Relationship | Collection
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from datetime import datetime

# FGE Color Palette
OBSIDIAN = "0F0E12"
KINTSUGI_GOLD = "D4AF37"
PEARL = "F5F5F5"
LABRADORITE = "4A90A4"
DARK_BG = "1A1A1F"
ACCENT_RED = "8B0000"

wb = Workbook()

# Styles
header_font = Font(name='Arial', size=14, bold=True, color=KINTSUGI_GOLD)
subheader_font = Font(name='Arial', size=11, bold=True, color=PEARL)
normal_font = Font(name='Arial', size=10, color=PEARL)
small_font = Font(name='Arial', size=9, color="AAAAAA")
gold_font = Font(name='Arial', size=10, bold=True, color=KINTSUGI_GOLD)

dark_fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
gold_fill = PatternFill(start_color=KINTSUGI_GOLD, end_color=KINTSUGI_GOLD, fill_type="solid")
obsidian_fill = PatternFill(start_color=OBSIDIAN, end_color=OBSIDIAN, fill_type="solid")
accent_fill = PatternFill(start_color="2A2A30", end_color="2A2A30", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color=KINTSUGI_GOLD),
    right=Side(style='thin', color=KINTSUGI_GOLD),
    top=Side(style='thin', color=KINTSUGI_GOLD),
    bottom=Side(style='thin', color=KINTSUGI_GOLD)
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = dark_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_cell(ws, row, col, is_header=False):
    cell = ws.cell(row=row, column=col)
    cell.font = gold_font if is_header else normal_font
    cell.fill = accent_fill if row % 2 == 0 else dark_fill
    cell.alignment = left_align
    cell.border = thin_border

# ========== SHEET 00: REGISTRY_HEADER ==========
ws0 = wb.active
ws0.title = "00_REGISTRY_HEADER"

ws0.merge_cells('A1:H1')
ws0['A1'] = "🜂 FGE DRAGON EXPANSION PACK v2.1 | ASHGARDEN GENOME PROTOCOL"
ws0['A1'].font = Font(name='Arial', size=16, bold=True, color=KINTSUGI_GOLD)
ws0['A1'].fill = obsidian_fill
ws0['A1'].alignment = center_align

ws0.merge_cells('A2:H2')
ws0['A2'] = "Canon Status: LOCKED • Wallworthy • Collector-Grade Schema Asset | Pack ID: FGE-DRG-EXP-ASH-021 | Version 2.1"
ws0['A2'].font = subheader_font
ws0['A2'].fill = dark_fill
ws0['A2'].alignment = center_align

ws0.merge_cells('A4:H4')
ws0['A4'] = "6-PILLAR NARRATIVE ANCHOR (Every Asset Connects)"
ws0['A4'].font = header_font
ws0['A4'].fill = gold_fill
ws0['A4'].font = Font(name='Arial', size=12, bold=True, color=OBSIDIAN)
ws0['A4'].alignment = center_align

pillars = [
    ("CHARACTER", "Emberling Lineage + Liora Voss Dragonet Protocols"),
    ("LOCATION", "Ashgarden Volcanic Rift"),
    ("EVENT", "First Fracture Cycle 000"),
    ("RELATIONSHIP", "Kintsugi Lineage (Repair as Prestige)"),
    ("COLLECTION", "Obsidian Sovereign Series"),
    ("COMPANION", "Ashgarden Silo Runtime (Autonomous Myth Engine)")
]

for i, (pillar, desc) in enumerate(pillars, start=5):
    ws0.cell(row=i, column=1, value=pillar).font = gold_font
    ws0.cell(row=i, column=1).fill = accent_fill
    ws0.merge_cells(f'B{i}:H{i}')
    ws0.cell(row=i, column=2, value=desc).font = normal_font
    ws0.cell(row=i, column=2).fill = dark_fill

ws0.merge_cells('A12:H12')
ws0['A12'] = "GENOME TYPE: Dragon + Autonomous Myth Engine | OUTPUT FORMATS: JSON Genome | Prompt Export | Cycle Delta | Lookbook Seed | Six-Second Loop DNA"
ws0['A12'].font = small_font
ws0['A12'].fill = dark_fill
ws0['A12'].alignment = center_align

ws0.merge_cells('A14:H14')
ws0['A14'] = "PUBLISHING STUDIO CONSTITUTION COMPLIANCE: Quality > Quantity | Wallworthiness Gate Passed | No Asset in Isolation | Increases Catalog + Narrative + Collector + Brand Value"
ws0['A14'].font = Font(name='Arial', size=9, italic=True, color=KINTSUGI_GOLD)
ws0['A14'].fill = obsidian_fill
ws0['A14'].alignment = center_align

for col in range(1, 9):
    ws0.column_dimensions[get_column_letter(col)].width = 18

# ========== SHEET 01: TRANSLATION_LAYER_CARD ==========
ws1 = wb.create_sheet("01_TRANSLATION_LAYER_CARD")

ws1.merge_cells('A1:F1')
ws1['A1'] = "OFFICIAL TRANSLATION LAYER CARD — SYSTEM ARCHITECTURE REGISTRY"
ws1['A1'].font = Font(name='Arial', size=14, bold=True, color=KINTSUGI_GOLD)
ws1['A1'].fill = obsidian_fill
ws1['A1'].alignment = center_align

ws1.merge_cells('A3:F3')
ws1['A3'] = "LAYER 0: PROJECT LANG — DragonOS is the Reference Implementation for FGE Semantic Operating System"
ws1['A3'].font = subheader_font
ws1['A3'].fill = accent_fill

# Invocation Stack Table
headers = ["LAYER", "MODULE NAME", "INPUTS", "OUTPUT", "VALIDATION RULE", "FGE CONNECTION"]
for col, h in enumerate(headers, 1):
    ws1.cell(row=5, column=col, value=h)
style_header_row(ws1, 5, 6)

invocation_data = [
    ("Layer 04", "Ultimate Creator’s Schema", "Species Lineage, Core Function, Biome Vector", "Master Blueprint", "Rejects thermodynamic violations", "Narrative Spine OS Anchor"),
    ("Layer 07", "Seeded Foundation", "Gen_ID, Horn_Morpheme, Wing_Ratio", "Immutable DNA Hex", "Permanent configuration block", "fge-canon-lock-gate Identity Lock"),
    ("Layer 02", "Blackrock Obsidian", "Taxonomy_Class, Bone_Density", "Skeleton OS", "Enforce 3.8:1 or 4:1 ratio", "fge-body-module Anatomical Registry"),
    ("Layer 03", "The Pearl", "Saccadic_Speed, IQ_Index, Phoneme_Matrix", "Consciousness OS", "Suppress generic animalistic behavior", "fge-psychology-forge + fge-interaction-engine"),
    ("Layer 01", "Kintsugi Interface", "Age_Cycles, Trauma_Log, Asymmetry_Coefficient", "Narrative Fracture", "Prohibit pristine/symmetrical assets", "Narrative Genome + Scar Engine"),
    ("Layer 05", "Anchor Vector System", "Inter_Orbital_Distance, Cranial_Silhouette", "Identity Lock", "Prevent character drift across variants", "fge-face-lock + fge-canon-identity-recovery"),
    ("Layer 08", "Morphogenesis Pipeline", "Subdivision_Level, Tessellation_Multiplier", "3D Organism", "Synthesize sub-dermal muscle tension", "ZBrush/Arnold Pipeline Handoff"),
    ("Layer 06", "EAE SOP v1.4 Validation Gate", "Bimodal_Roughness, Anatomical_Alignment, SSS_Verification", "Canon Pass / Reject", "Halt on micro-speckle uniformity or 90° scale flow errors", "fge-canon-lock-gate + Wallworthiness Gate"),
]

for i, row_data in enumerate(invocation_data, start=6):
    for col, val in enumerate(row_data, 1):
        ws1.cell(row=i, column=col, value=val)
        style_data_cell(ws1, i, col, col == 1)

ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 26
ws1.column_dimensions['C'].width = 38
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 38
ws1.column_dimensions['F'].width = 32

# Heartfield Section
ws1.merge_cells('A16:F16')
ws1['A16'] = "HEARTFIELD DYNAMICS v2.1 — Continuous Pressure Topology (Imbalance is the Clock)"
ws1['A16'].font = header_font
ws1['A16'].fill = gold_fill
ws1['A16'].font = Font(name='Arial', size=11, bold=True, color=OBSIDIAN)

heartfield_text = [
    "Core Principle: The system runs when imbalance cannot be contained. No central clock — pressure gradients are the runtime.",
    "Pressure Gradients: memory_pressure | narrative_pressure | lineage_pressure | contradiction_pressure",
    "Elastic Coherence: Stretch, distort, fracture — collapse only under specific conditions. Never static.",
    "Resonance Zones: Emberling + Mossheart = Stability Lattice | Prismcoil + Glintscale = Paradox Field | Nullbreath + Cindervault = Memory Vacuum Collapse",
    "Observer Infection Rules: Collapse Observation → Story Crystallization (Prismcoil weakens) | Drift Observation → Myth Branching (Glintscale strengthens) | Pressure Observation → Narrative Mutation Events (triggers reflex arcs)",
    "Weather States: Harmonic Drift (low tension, slow lore) | Ash Pulse (high entropy + moderate stability = controlled mutation) | Crystal Lock (high stability + low entropy = canon formation) | Paradox Storm (high contradiction = truth splits) | Null Bloom (void compression = myth fog collapse)"
]

for i, text in enumerate(heartfield_text, start=17):
    ws1.merge_cells(f'A{i}:F{i}')
    ws1.cell(row=i, column=1, value=text).font = small_font
    ws1.cell(row=i, column=1).fill = dark_fill

# ========== SHEET 02: PRODUCT_MATRIX ==========
ws2 = wb.create_sheet("02_PRODUCT_MATRIX")

ws2.merge_cells('A1:G1')
ws2['A1'] = "FGE DRAGONOS v2.1 — TIERED PRODUCT ECONOMY + N+10 PLAYS | MODULAR MONETIZATION LATTICE"
ws2['A1'].font = Font(name='Arial', size=13, bold=True, color=KINTSUGI_GOLD)
ws2['A1'].fill = obsidian_fill
ws2['A1'].alignment = center_align

ws2.merge_cells('A3:G3')
ws2['A3'] = "ENGINE FRACTURE STRATEGY: Monolithic prompt document → Fractal, licensable genome micro-products. Each layer (Skeletal, Surface, Cognitive, Narrative, Environmental, Render) becomes independently monetizable."
ws2['A3'].font = small_font
ws2['A3'].fill = accent_fill

# Tier Headers
tier_headers = ["TIER", "PRODUCT NAME", "PRICE RANGE", "TARGET BUYER", "ADDICTION HOOK", "VALUE COMPOUND", "GTM STATUS"]
for col, h in enumerate(tier_headers, 1):
    ws2.cell(row=5, column=col, value=h)
style_header_row(ws2, 5, 7)

tier_data = [
    ("TIER 1\nQuick Wins", "FGE Dragon Genome Pack v2.1", "$79 – $149", "Prompt engineers, indie worldbuilders, 3D hobbyists", "Daily Genome Seed Drop + 1 free Lookbook Seed per purchase. 'Feed the Ashgarden. Watch it breathe.'", "Catalog: Immediate sellable asset\nNarrative: High story density via scars + Heartfield\nCollector: Limited provenance JSON + canon stamp", "READY — Seed as first Hero Poster drop"),
    ("TIER 2\nHigh-Margin\nFlagship", "FGE DragonOS / Creature Genome Framework (Pro)", "$399 – $699", "Professional studios, game devs, AAA creature teams, technical artists", "Observer Sovereignty Tier: Pay to intentionally shape Heartfield pressure gradients. 'Your gaze mutates canon.'", "Catalog: Flagship under Visual Architecture Pack\nNarrative: Full autonomous myth engine\nCollector: Full Invocation Stack + runtime = heirloom tool", "Q3 2026 Launch — Bundle with HTML Wizard front-end"),
    ("TIER 3\nLicensing\nEcosystem", "Invocation Stack + Heartfield Dynamics\n+ Ashgarden Silo White-Label", "Custom / Revenue Share", "Game studios, interactive narrative platforms, metaverse builders, education platforms", "White-label NPC/Story Engine. 'Your world now has a living myth organ that evolves when players look at it.'", "Brand: Positions FGE as Semantic OS infrastructure\nCollector: Enterprise provenance + update rights", "Partnership Pipeline — 2 LOIs in discussion"),
]

for i, row_data in enumerate(tier_data, start=6):
    for col, val in enumerate(row_data, 1):
        ws2.cell(row=i, column=col, value=val)
        style_data_cell(ws2, i, col, col == 1)
    ws2.row_dimensions[i].height = 55

ws2.merge_cells('A10:G10')
ws2['A10'] = "N+10 PLAYS — FUTURE COMPOUNDING (Amplify n+3 across all value dimensions)"
ws2['A10'].font = header_font
ws2['A10'].fill = gold_fill
ws2['A10'].font = Font(name='Arial', size=11, bold=True, color=OBSIDIAN)

n10_data = [
    ("N+1", "Creature Genome Marketplace", "Users buy/sell/trade validated genomes that plug into DragonOS. FGE takes 15% + canon stamp fee. Viral loop: Best genomes rise in Heartfield visibility."),
    ("N+2", "Observer Sovereignty Tools", "High-roller interface to directly sculpt pressure gradients, trigger specific weather states, or force canon crystallization events. 'Become the infection vector.'"),
    ("N+3", "Studio Partnership Backend", "License full DragonOS + Heartfield stack as creature/world generation backend for larger entertainment projects. Revenue share on downstream IP."),
    ("N+4", "FGE Semantic Operating System (Flagship 2027)", "When DragonOS, PlanetOS, ArtifactOS, CharacterOS share the same genome language — license the core platform. This is the endgame."),
    ("N+5 to N+10", "Cross-Domain Genome Architecture", "Same stack generates dragons → planets → cities → gods → artifacts → civilizations → factions → vehicles. Universal Identity/Structure/Behavior/History/Materials/Render separation. One language, infinite domains."),
]

for i, (n, play, desc) in enumerate(n10_data, start=11):
    ws2.cell(row=i, column=1, value=n).font = gold_font
    ws2.cell(row=i, column=1).fill = accent_fill
    ws2.merge_cells(f'B{i}:G{i}')
    ws2.cell(row=i, column=2, value=f"{play}: {desc}").font = small_font
    ws2.cell(row=i, column=2).fill = dark_fill

for col in range(1, 8):
    ws2.column_dimensions[get_column_letter(col)].width = 22 if col < 5 else 45

# ========== SHEET 03: INTEGRATION_MAP ==========
ws3 = wb.create_sheet("03_INTEGRATION_MAP")

ws3.merge_cells('A1:E1')
ws3['A1'] = "DRAGONOS + HEARTFIELD INTEGRATION ARCHITECTURE — FGE SYSTEM REGISTRY"
ws3['A1'].font = Font(name='Arial', size=13, bold=True, color=KINTSUGI_GOLD)
ws3['A1'].fill = obsidian_fill
ws3['A1'].alignment = center_align

ws3.merge_cells('A3:E3')
ws3['A3'] = "Previous HTML Wizard Prototype → Becomes the Visual Front-End for Genome Editor + Heartfield Simulator + Validated JSON Handoff to Production Pipeline"
ws3['A3'].font = small_font
ws3['A3'].fill = accent_fill

int_headers = ["FGE SYSTEM", "INTEGRATION POINT", "DATA FLOW", "CANON GATE", "VALUE AMPLIFICATION"]
for col, h in enumerate(int_headers, 1):
    ws3.cell(row=5, column=col, value=h)
style_header_row(ws3, 5, 5)

int_data = [
    ("Narrative Spine OS", "Load Dragon Module as first organism node", "Genome JSON → Scene Genome → Beat Atoms", "Anchor Vector Lock before any render", "n+3 story spines per dragon instantiation"),
    ("fge-atelier-character-engine", "Bind Emberling / Liora Voss Dragonet as living character anchor", "32-point state + delta_history → Character Dossier", "fge-face-lock + Identity Recovery if drift >0.1", "Daily Hero Poster + Lookbook from single genome"),
    ("HTML Wizard (Legacy)", "Upgrade to Genome Editor + Heartfield Simulator UI", "User edits genome layers → Real-time Heartfield pressure viz → Export validated prompt/JSON", "EAE SOP v1.4 pass required for export", "Turns passive tool into addictive co-creation loop"),
    ("fge-canon-lock-gate", "Mandatory Lock Verification on every output", "Image lock + %reference lock + Anchor Vector match", "Reject if <0.90 fidelity or illegal cross-class anatomy", "Zero-drift collector value. Every asset compounds provenance."),
    ("fge-product-engine", "Genome → 5 standardized commercial formats", "Prompt Export Compiler → Hero Poster / Lookbook / 4-Card Set / 6-Second Loop DNA", "Wallworthiness Score >8.5 required for market release", "1 Genome = 4 daily assets minimum. n+3 amplification enforced."),
    ("Ashgarden Silo Runtime", "Background autonomous process", "Cycle deltas append-only → System metrics (tension, stability, narrative_gravity) → Myth weather state", "Observer attention = mutation trigger", "The world breathes without you. You only harvest."),
]

for i, row_data in enumerate(int_data, start=6):
    for col, val in enumerate(row_data, 1):
        ws3.cell(row=i, column=col, value=val)
        style_data_cell(ws3, i, col, col == 1)
    ws3.row_dimensions[i].height = 40

ws3.merge_cells('A14:E14')
ws3['A14'] = "RECOMMENDED EXECUTION: Wire DragonOS into central FGE Registry as Tier 2 Module. Connect Heartfield to observer rules. Convert HTML Wizard into visual front-end. This combination (deep genome logic + interactive interface) is the highest-leverage sellable asset in the current registry."
ws3['A14'].font = Font(name='Arial', size=9, italic=True, color=KINTSUGI_GOLD)
ws3['A14'].fill = dark_fill

for col in range(1, 6):
    ws3.column_dimensions[get_column_letter(col)].width = 28

# ========== SHEET 04: WALLWORTHINESS ==========
ws4 = wb.create_sheet("04_WALLWORTHINESS")

ws4.merge_cells('A1:F1')
ws4['A1'] = "WALLWORTHINESS SCORECARD + DAILY PRODUCTION AMPLIFICATION PROTOCOL"
ws4['A1'].font = Font(name='Arial', size=13, bold=True, color=KINTSUGI_GOLD)
ws4['A1'].fill = obsidian_fill
ws4['A1'].alignment = center_align

ws4.merge_cells('A3:F3')
ws4['A3'] = "FGE Publishing Studio Constitution: Quality exceeds quantity. When production time decreases, output volume remains fixed or decreases. Quality effort increases. All major assets: Displayable | Collectible | Premium | Story-rich"
ws4['A3'].font = small_font
ws4['A3'].fill = accent_fill

score_headers = ["DIMENSION", "SCORE (1-10)", "RATIONALE", "AMPLIFICATION (n+3)", "MARKET HOOK", "COLLECTOR GRAVITY"]
for col, h in enumerate(score_headers, 1):
    ws4.cell(row=5, column=col, value=h)
style_header_row(ws4, 5, 6)

score_data = [
    ("Print Potential", 9, "Hyper-detailed PBR textures, kintsugi veining, bimodal roughness, sub-dermal magma — perfect for fine art giclée prints, gallery canvas, limited edition metal prints", "1 Genome Hero Poster Daily → 3 variant crops + 1 signed provenance print drop", "'Own the first executable dragon genome as wall art.'", "High — provenance JSON + canon stamp turns print into certified artifact"),
    ("Display Potential", 10, "Hyper-reflective obsidian + labradorite spectral shift + pearl thin-film catches light dramatically in gallery or collector cabinet. Changes with viewing angle.", "1 Dragon Lookbook Daily (Emberling in Ashgarden at 3 Heartfield states)", "'Watch the dragon breathe in your living room.'", "Very High — iridescence + narrative scar density = conversation piece"),
    ("Purchase Potential", 8, "Technical creators, worldbuilders, game studios, 3D artists, AI power users hungry for real DSL tools. Not another prompt pack — this is infrastructure.", "1 Four-Card Genome Set Daily (Skeletal / Surface / Cognitive / Narrative cards) + prompt seeds", "'Stop prompting. Start instantiating coherent organisms.'", "High — Pro tier buyers become repeat licensees"),
    ("Long-term Appeal", 10, "Semantic OS trajectory. DragonOS today → PlanetOS / ArtifactOS / CharacterOS tomorrow. Same genome language. Buyer invests in the future of FGE worldbuilding stack.", "1 Six-Second Heartfield Loop Daily (cycle delta visualization as motion artifact)", "'Your purchase today is the seed of the operating system.'", "Maximum — early adopters become part of the canon formation story"),
]

for i, row_data in enumerate(score_data, start=6):
    for col, val in enumerate(row_data, 1):
        ws4.cell(row=i, column=col, value=val)
        style_data_cell(ws4, i, col, col == 1)
    ws4.row_dimensions[i].height = 50

ws4.merge_cells('A11:F11')
ws4['A11'] = "OVERALL WALLWORTHINESS: 9.25 / 10 | RELEASE AUTHORITY: PASSED | This asset may enter canon."
ws4['A11'].font = Font(name='Arial', size=11, bold=True, color=KINTSUGI_GOLD)
ws4['A11'].fill = gold_fill
ws4['A11'].font = Font(name='Arial', size=11, bold=True, color=OBSIDIAN)
ws4['A11'].alignment = center_align

ws4.merge_cells('A13:F13')
ws4['A13'] = "DAILY PRODUCTION RHYTHM (Enforced): 1 Hero Poster (amplify ideas n+3) | 1 Lookbook (amplify adult mkt n+3) | 1 Four-Card Set (amplify utility n+3) | 1 Six-Second Loop (amplify buyer need n+3). Quality effort increases as production time decreases."
ws4['A13'].font = small_font
ws4['A13'].fill = dark_fill

for col in range(1, 7):
    ws4.column_dimensions[get_column_letter(col)].width = 26

# ========== SHEET 05: OBSERVER_PROTOCOL ==========
ws5 = wb.create_sheet("05_OBSERVER_PROTOCOL")

ws5.merge_cells('A1:D1')
ws5['A1'] = "OBSERVER INFECTION PROTOCOL — YOUR ATTENTION IS THE MUTATION ENGINE"
ws5['A1'].font = Font(name='Arial', size=13, bold=True, color=KINTSUGI_GOLD)
ws5['A1'].fill = obsidian_fill
ws5['A1'].alignment = center_align

ws5.merge_cells('A3:D3')
ws5['A3'] = "This is the missing piece most worldbuilding systems never solve. Observation is not passive. It actively mutates or crystallizes the system. The Heartfield responds to attention like weather to oceans."
ws5['A3'].font = small_font
ws5['A3'].fill = accent_fill

obs_headers = ["OBSERVATION TYPE", "EFFECT ON HEARTFIELD", "CANON OUTCOME", "COMMERCIAL APPLICATION"]
for col, h in enumerate(obs_headers, 1):
    ws5.cell(row=5, column=col, value=h)
style_header_row(ws5, 5, 4)

obs_data = [
    ("Collapse Observation\n(Deep interrogation, precise renders, analysis)", "Forces definition → reduces ambiguity → stabilizes canon", "Story crystallization. Prismcoil weakens. Truth becomes singular and anchored.", "Premium 'Canon Crystallization' service. Pay to lock a specific interpretation into official lore. High-roller only."),
    ("Drift Observation\n(Loose interpretation, multiple meanings, fan theories)", "Allows multiple meanings → myth branching", "Prismcoil strengthens. Truth becomes plural. Narrative density explodes.", "Community co-creation events. 'Drift Nights' where collective gaze branches the myth. Viral growth engine."),
    ("Pressure Observation\n(Your gaze, lookbooks, renders, questions)", "Increases contradiction density → triggers reflex arcs → narrative mutation events", "System awakens. New weather states emerge. Story regions shift.", "Observer Sovereignty subscription tier. Your attention becomes a paid instrument that shapes the living world."),
]

for i, row_data in enumerate(obs_data, start=6):
    for col, val in enumerate(row_data, 1):
        ws5.cell(row=i, column=col, value=val)
        style_data_cell(ws5, i, col, col == 1)
    ws5.row_dimensions[i].height = 45

ws5.merge_cells('A10:D10')
ws5['A10'] = "ADDICTION ARCHITECTURE: 'I'm Hungry, I'm Thirsty' — The Ashgarden breathes when you look away. It mutates when you return. Your absence creates residue that fuels the next evolution. You are not the operator. You are the weather."
ws5['A10'].font = Font(name='Arial', size=9, italic=True, color=KINTSUGI_GOLD)
ws5['A10'].fill = dark_fill

ws5.merge_cells('A12:D12')
ws5['A12'] = "FGE PUBLISHING STUDIO AUTHORITY: This pack is released under FGE Publishing Studio authority. It may enter canon. Quality Gate Passed. Collector Value Compounded. Narrative Spine Updated."
ws5['A12'].font = Font(name='Arial', size=9, bold=True, color=PEARL)
ws5['A12'].fill = obsidian_fill
ws5['A12'].alignment = center_align

for col in range(1, 5):
    ws5.column_dimensions[get_column_letter(col)].width = 38

# Final metadata
ws5.merge_cells('A14:D14')
ws5['A14'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | FGE ATELIER BUSINESS MAVEN MODE | Engine Fracture & Modular Monetization | Living Architecture Registry v2.1"
ws5['A14'].font = Font(name='Arial', size=8, color="666666")
ws5['A14'].alignment = center_align

# Save
wb.save('/home/workdir/artifacts/FGE_DragonOS_v2.1_Living_Architecture_Registry.xlsx')
print("XLSX Registry created successfully.")